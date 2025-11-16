"""
Exploratory Data Analysis for LinkedIn Excel Data
This script analyzes the structure of the LinkedIn analytics export file
to understand the data we're working with.
"""

import polars as pl
from pathlib import Path


def explore_excel_file(file_path: str) -> None:
    """
    Explore the structure of the LinkedIn Excel file.

    Args:
        file_path: Path to the Excel file to analyze
    """
    print(f"📊 Exploring: {file_path}\n")
    print("=" * 80)

    # Read all sheets as a dictionary
    print("🔍 Reading all sheets from the Excel file...\n")
    try:
        # Use sheet_id=0 to get all sheets as a dict
        try:
            sheets = pl.read_excel(
                file_path,
                sheet_id=0,
                engine="openpyxl"
            )
        except TypeError as te:
            # Polars may raise TypeError when column names are not strings
            # Fallback: use openpyxl to load values, coerce headers to strings,
            # then convert into polars DataFrames per sheet.
            from openpyxl import load_workbook

            print("⚠️  Polars read_excel failed with TypeError, falling back to openpyxl...\n")
            wb = load_workbook(file_path, data_only=True, read_only=True)
            sheets = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.values)
                print(f"  -> Raw rows in sheet '{sheet_name}': {len(rows)}")
                if rows:
                    # show a small preview of the raw rows to understand where headers/data live
                    preview = rows[:10]
                    print("     First rows preview (up to 10):")
                    for r in preview:
                        print(f"      {r}")
                if not rows:
                    sheets[sheet_name] = pl.DataFrame()
                    continue

                # First row is header if it contains at least one non-None value
                header_row = rows[0]
                header = []
                for i, h in enumerate(header_row):
                    if h is None:
                        header.append(f"column_{i+1}")
                    else:
                        header.append(str(h))

                data_rows = rows[1:]
                if data_rows:
                    try:
                        df = pl.DataFrame(data_rows, schema=header)
                    except Exception:
                        # If creating DataFrame with schema fails, build via dict
                        cols = {h: [r[idx] if idx < len(r) else None for r in data_rows]
                                for idx, h in enumerate(header)}
                        df = pl.DataFrame(cols)
                else:
                    # Empty sheet with header only
                    df = pl.DataFrame({h: [] for h in header})

                sheets[sheet_name] = df

        # Sheets is now a dict of {sheet_name: DataFrame}
        print(f"✅ Found {len(sheets)} sheet(s):\n")

        for sheet_name, df in sheets.items():
            print(f"\n{'=' * 80}")
            print(f"📄 Sheet: {sheet_name}")
            print(f"{'=' * 80}")
            print(f"Shape: {df.shape[0]} rows × {df.shape[1]} columns")

            print(f"\n📋 Columns ({len(df.columns)}):")
            for i, col in enumerate(df.columns, 1):
                print(f"  {i:2d}. {col}")

            print(f"\n📊 Column Data Types:")
            for col_name, col_type in df.schema.items():
                print(f"  • {col_name}: {col_type}")

            print(f"\n📈 First 5 rows:")
            print(df.head(5))

            print(f"\n📉 Data Summary:")
            print(f"  • Total rows: {df.shape[0]}")
            print(f"  • Total columns: {df.shape[1]}")

            # Show null counts
            null_counts = df.null_count()
            # null_counts is a one-row DataFrame; convert to a dict of {col: count}
            try:
                null_row = null_counts.row(0, named=True)
            except Exception:
                null_row = {}

            if any((v > 0) for v in null_row.values()):
                print(f"\n⚠️  Null Value Counts:")
                for col, count in null_row.items():
                    if count > 0:
                        print(f"  • {col}: {count}")

            # Show basic statistics for numeric columns
            numeric_cols = [col for col, dtype in df.schema.items()
                            if dtype in (pl.Int32, pl.Int64, pl.Float32, pl.Float64)]
            if numeric_cols:
                print(f"\n📊 Numeric Columns Statistics:")
                print(df.select(numeric_cols).describe())

    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    # Path to the LinkedIn analytics export
    excel_file = Path(__file__).parent.parent.parent / "test-data" / "Content_2024-11-11_2025-11-10_ShifraWilliams.xlsx"

    if excel_file.exists():
        explore_excel_file(str(excel_file))
    else:
        print(f"❌ File not found: {excel_file}")
        print("\nSearching for Excel files in test-data...")
        test_data_dir = Path(__file__).parent.parent.parent / "test-data"
        if test_data_dir.exists():
            excel_files = list(test_data_dir.glob("*.xlsx"))
            if excel_files:
                print(f"Found: {excel_files}")
            else:
                print("No .xlsx files found")
