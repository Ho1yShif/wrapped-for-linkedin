"""
Parser for LinkedIn Excel export sheets.
Handles extraction of data from specific sheets with their unique structures.
"""

from typing import Dict, Any, Optional
from openpyxl import load_workbook
import io


def parse_discovery_sheet(file_content: bytes) -> Dict[str, Any]:
    """
    Parse the DISCOVERY sheet from LinkedIn Excel export.

    Expected structure:
    Row 1: "Overall Performance" | "11/11/2024 - 11/10/2025"
    Row 2: "Impressions" | 857000
    Row 3: "Members reached" | 297771

    Args:
        file_content: Binary content of Excel file

    Returns:
        Dictionary with keys: "Overall Performance", "Impressions", "Members reached"
    """
    try:
        wb = load_workbook(io.BytesIO(file_content), data_only=True)

        if "DISCOVERY" not in wb.sheetnames:
            raise ValueError("DISCOVERY sheet not found in Excel file")

        ws = wb["DISCOVERY"]

        discovery_data = {}

        # Read first 3 rows (label in column A, value in column B)
        for row_idx in range(1, 4):
            cell_label = ws.cell(row_idx, 1).value
            cell_value = ws.cell(row_idx, 2).value

            if cell_label is not None:
                discovery_data[str(cell_label)] = cell_value

        return discovery_data

    except Exception as e:
        raise ValueError(f"Failed to parse DISCOVERY sheet: {str(e)}")
