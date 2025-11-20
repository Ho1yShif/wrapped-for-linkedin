"""
Parser for TOP POSTS sheet from LinkedIn Excel export.
Extracts posts with engagement metrics and URLs.
"""

from typing import Dict, Any, List, Optional
from openpyxl import load_workbook
import io
from datetime import datetime


def parse_top_posts_sheet(file_content: bytes) -> List[Dict[str, Any]]:
    """
    Parse the TOP POSTS sheet from LinkedIn Excel export.

    Expected structure:
    Row 1: Header note
    Row 2: Empty
    Row 3: Column headers - "Post URL", "Post publish date", "Engagements", ..., "Post URL", "Post publish date", "Impressions"
    Row 4+: Data rows with post information

    The sheet has two views of the data:
    - Left side (columns A-C): Post URL, Publish Date, Engagements
    - Right side (columns E-G): Post URL, Publish Date, Impressions

    Args:
        file_content: Binary content of Excel file

    Returns:
        List of dictionaries containing post data with keys:
        - url: str
        - publish_date: str
        - engagements: float
        - impressions: float (optional)
    """
    try:
        wb = load_workbook(io.BytesIO(file_content), data_only=True)

        if "TOP POSTS" not in wb.sheetnames:
            raise ValueError("TOP POSTS sheet not found in Excel file")

        ws = wb["TOP POSTS"]

        posts = {}  # Use dict to merge engagement and impression data by URL

        # Row 3 contains headers, data starts from row 4
        # Parse engagement data from columns A-C
        for row_idx in range(4, ws.max_row + 1):
            url = ws.cell(row_idx, 1).value
            publish_date = ws.cell(row_idx, 2).value
            engagements = ws.cell(row_idx, 3).value

            # Stop if we hit empty rows
            if url is None or publish_date is None or engagements is None:
                # Check if all first 3 columns are empty (end of data)
                if (ws.cell(row_idx, 1).value is None and
                    ws.cell(row_idx, 2).value is None and
                    ws.cell(row_idx, 3).value is None):
                    break

            # Skip empty rows
            if url is None or publish_date is None:
                continue

            # Convert engagements to float
            try:
                engagements_count = float(engagements) if engagements else 0
            except (ValueError, TypeError):
                engagements_count = 0

            url_str = str(url)
            posts[url_str] = {
                "url": url_str,
                "publish_date": str(publish_date),
                "engagements": engagements_count,
                "impressions": 0
            }

        # Parse impression data from columns E and G and merge with engagement data
        # In openpyxl, columns are 1-indexed: A=1, B=2, C=3, D=4, E=5, F=6, G=7
        for row_idx in range(4, ws.max_row + 1):
            url = ws.cell(row_idx, 5).value  # Column E
            impressions = ws.cell(row_idx, 7).value  # Column G

            if url is None or impressions is None:
                if (ws.cell(row_idx, 5).value is None and
                    ws.cell(row_idx, 7).value is None):
                    break
                continue

            # Convert impressions to float
            try:
                impressions_count = float(impressions) if impressions else 0
            except (ValueError, TypeError):
                impressions_count = 0

            url_str = str(url)
            if url_str in posts:
                posts[url_str]["impressions"] = impressions_count
            else:
                # If URL not found in engagement data, still add it
                posts[url_str] = {
                    "url": url_str,
                    "publish_date": "",
                    "engagements": 0,
                    "impressions": impressions_count
                }

        # Convert dict to sorted list
        posts_list = list(posts.values())

        # Sort by engagements descending
        posts_list.sort(key=lambda x: x["engagements"], reverse=True)

        return posts_list

    except Exception as e:
        raise ValueError(f"Failed to parse TOP POSTS sheet: {str(e)}")

def get_top_posts(posts: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Get the top 6 posts by engagement count.

    Args:
        posts: List of post dictionaries

    Returns:
        Top 6 posts with rank added
    """
    top_posts = posts[:6]
    for idx, post in enumerate(top_posts, 1):
        post["rank"] = idx
    return top_posts
